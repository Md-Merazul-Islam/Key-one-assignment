import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
import time


def get_suggestions(keyword):
    driver = webdriver.Chrome()

    try:
        # Navigate to Google
        driver.get("https://www.google.com")
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)
        time.sleep(2)

        # Collect suggestions
        suggestions = driver.find_elements(
            By.CSS_SELECTOR, 'div[role="presentation"] span')
        longest = ""
        shortest = None

        for suggestion in suggestions:
            text = suggestion.text.strip()
            if text:
                if len(text) > len(longest):
                    longest = text
                if shortest is None or len(text) < len(shortest):
                    shortest = text

        print(f"Keyword: {keyword}")
        print("Longest Suggestion: ", longest)
        print("Shortest Suggestion: ", shortest)

        return longest, shortest

    finally:
        driver.quit()


def create_or_update_sheet(file_path, day, keyword, longest, shortest):
    try:
        # Load or create Excel file
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        # Check if the day's sheet exists; if not, create it
        if day not in wb.sheetnames:
            sheet = wb.create_sheet(title=day)
            sheet.append(["Keyword", "Longest Suggestion",
                         "Shortest Suggestion"])
        else:
            sheet = wb[day]

        # Append the keyword data to the sheet
        sheet.append([keyword, longest, shortest])

        # Save the workbook
        wb.save(file_path)
        print(f"Updated sheet for {day} with keyword: {keyword}")
    except Exception as e:
        print(f"Error while updating Excel sheet: {e}")


if __name__ == "__main__":
    excel_file = "keywords.xlsx"
    today = datetime.now().strftime("%A")

    # Load the workbook and get the keywords for today
    try:
        wb = openpyxl.load_workbook(excel_file)
        if today not in wb.sheetnames:
            print(f"Sheet for {today} does not exist. Exiting the script.")
        else:
            sheet = wb[today]
            keywords = [sheet.cell(row=row, column=1).value for row in range(
                2, sheet.max_row + 1)]

            # Process each keyword
            for keyword in keywords:
                if keyword:
                    print(f"Processing keyword: {keyword}")
                    longest, shortest = get_suggestions(keyword)
                    create_or_update_sheet(
                        excel_file, today, keyword, longest, shortest)

            print("Excel file has been updated.")
    except Exception as e:
        print(f"Error while processing the Excel file: {e}")

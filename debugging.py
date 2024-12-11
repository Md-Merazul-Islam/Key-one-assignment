from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from datetime import datetime

def initialize_driver():
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    return driver


def get_suggestions(driver, keyword):
    try:
        driver.get("https://www.google.com/")
        print("Google page loaded.")
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(keyword)
        print(f"Keyword entered: {keyword}")


        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'ul.erkvQe li span'))
        )
        print("Suggestions loaded.")

        time.sleep(2)  


        search_box.send_keys(Keys.RETURN)
        print("Search triggered.")

        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'ul.erkvQe li span'))
        )
        print("Suggestions are now available after pressing return.")

        suggestions = driver.find_elements(By.CSS_SELECTOR, 'ul.erkvQe li span')


        if suggestions:
            suggestions_text = [suggestion.text for suggestion in suggestions]
            longest = max(suggestions_text, key=len)
            shortest = min(suggestions_text, key=len)
            print(f"Longest suggestion: {longest}, Shortest suggestion: {shortest}")
        else:
            print("No suggestions found.")
            longest, shortest = "", ""

    except Exception as e:
        print(f"Error during Google search: {e}")
        longest, shortest = "", ""
    
    return longest, shortest


def update_excel(file_path, day, keyword, longest, shortest):
    try:
        wb = openpyxl.load_workbook(file_path)

        if day not in wb.sheetnames:
            print(f"Sheet for {day} does not exist. Skipping update for today.")
            return

        sheet = wb[day] 


        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == keyword:
                sheet.cell(row=row, column=2).value = longest
                sheet.cell(row=row, column=3).value = shortest
                break

        wb.save(file_path)
    except Exception as e:
        print(f"Error while updating Excel file: {e}")


if __name__ == "__main__":

    excel_file = "keywords.xlsx"
    today = datetime.now().strftime("%A")
    try:
        wb = openpyxl.load_workbook(excel_file)


        if today not in wb.sheetnames:
            print(f"Sheet for {today} does not exist. Exiting the script.")
        else:
            sheet = wb[today]  
            keywords = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]

            driver = initialize_driver()

            for keyword in keywords:
                if keyword:  
                    print(f"Processing keyword: {keyword}")
                    longest, shortest = get_suggestions(driver, keyword)
                    print(f"Longest: {longest}, Shortest: {shortest}")
                    update_excel(excel_file, today, keyword, longest, shortest)

            print("Excel file has been updated.")
            driver.quit()  
    except Exception as e:
        print(f"Error while processing the Excel file: {e}")
